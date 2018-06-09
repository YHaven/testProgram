# coding=utf-8
import os
import sys
from tornado import gen
from tornado.web import authenticated
import tornado.escape 

from base import BaseHandler
from model.pager import Pager
from model.constants import Constants
from model.search_params.order_params import OrderSearchParams
from model.search_params.product_params import ProductSearchParams
from service.user_service import UserService
from service.home_service import HomeService
from openpyxl import load_workbook




reload(sys)   
sys.setdefaultencoding('utf8')

class OrderHandler(BaseHandler):
    @authenticated
    @gen.coroutine
    def get(self):
        pager = Pager(self)
        base_url = self.reverse_url('order.index')
        order_search_params = OrderSearchParams(self)
        pager = yield self.async_do(HomeService.page_order, self.db, pager, order_search_params)
        self.render("order/index.html", base_url=base_url,pager=pager,order_search_params=order_search_params)



class ProductList(BaseHandler):
    
    @gen.coroutine
    def get(self):
        pager = Pager(self)
        base_url = self.reverse_url('productList')
        product_search_params = ProductSearchParams(self)
        pager = yield self.async_do(HomeService.page_product, self.db, pager, product_search_params)

        self.render("product/productList.html", base_url=base_url,pager=pager,order_search_params=product_search_params)

class ProductListByOrder(BaseHandler):

    @gen.coroutine
    def get(self):
        pager = Pager(self)
        base_url = self.reverse_url('productListByOrder')
        product_search_params = ProductSearchParams(self)
        
        orderid = self.get_argument('id')
        ret = {'result': 'OK'}
        productList = []

        order = yield self.async_do(HomeService.get_order_by_id, self.db, orderid)
        if order is not None:

            # userMark = unicode(order.userMark_.encode("utf-8"), "gbk")
            userMark = order.userMark_
            classNameList__real = []
            otherNameList__real = []
            productNameList__real = []
            if(userMark.find('）')>0):
                # 先区分有多少类型 最后一个【米尊】保留着
                classNameList = userMark.split('）')   
                for className in classNameList:
                    if(className.find('（')>0):
                        className_real = className.split('（')[0]
                        # 排除第二个类型第一个字符是+
                        if(className_real.find('+') == 0):
                            className_real = className_real[1:]
                        print '类别：'+className_real
                        productNameAll = className.split('（')[1]
                        productName_real = productNameAll
                        
                        # 内有多个产品
                        if(productNameAll.find('+')>0):
                            productNameList = productNameAll.split('+')
                            for productName in productNameList:
                                productName_real = productName
                                otherName_real = ''
                                productNumber = 1
                                # 产品是否有别名？
                                if(productName_real.find('【')>0):
                                    otherName_real = productName_real.split('【')[1][:-1]
                                    productName_real = productName_real.split('【')[0]

                                # 检查是否有个数
                                if(productName_real.find('个')>0):
                                    productNumber = int(productName_real.split('个')[0])
                                    productName_real = productName_real.split('个')[1]
                                    
                                # print otherName_real
                                # print '产品Real：'+productName_real
                                # print productNumber
                                product_search_params.className = className_real
                                product_search_params.otherName = otherName_real
                                product_search_params.productName = productName_real
                                product = yield self.async_do(HomeService.page_product_byorder, self.db, product_search_params) 
                                if product is not None :
                                    productItem = dict(
                                        id = product.id,
                                        status_ = product.status_,
                                        picture_ = product.picture_,
                                        className_ = product.className_,
                                        productName_ = product.productName_,
                                        otherName_ = product.otherName_,
                                        price_ = product.price_,
                                        productNumber_ = productNumber,
                                        # createDate_ = product.createDate_,
                                    )
                                    productList.append(productItem)
                        # 只有一个产品
                        else:
                            otherName_real = ''
                            productNumber = 1
                            # 产品是否有别名？
                            if(productName_real.find('【')>0):
                                otherName = productName_real.split('【')[1][:-1]
                                productName_real = productName_real.split('【')[0]

                            # 检查是否有个数
                            if(productName_real.find('个')>0):
                                productNumber = int(productName_real.split('个')[0])
                                productName_real = productName_real.split('个')[1]
                                
                            # print otherName
                            # print '产品Real：'+productName_real
                            # print productNumber
                            product_search_params.className = className_real
                            product_search_params.otherName = otherName_real
                            product_search_params.productName = productName_real
                            product = yield self.async_do(HomeService.page_product_byorder, self.db, product_search_params) 
                            if product is not None :
                                productItem = dict(
                                    id = product.id,
                                    status_ = product.status_,
                                    picture_ = product.picture_,
                                    className_ = product.className_,
                                    productName_ = product.productName_,
                                    otherName_ = product.otherName_,
                                    price_ = product.price_,
                                    productNumber_ = productNumber,
                                    # createDate_ = product.createDate_,
                                )
                                productList.append(productItem)
            ret['status'] = 'true'
        else:
            ret['status'] = 'false'

        ret['productList'] = productList
        respon_json = tornado.escape.json_encode(ret)  
        self.write_json(respon_json)
        # self.render("productList.html", base_url=base_url,pager=pager,order_search_params=product_search_params)


class ProductAdd(BaseHandler):

    def get(self):
        pager = Pager(self)
        base_url = self.reverse_url('index')
        order_search_params = OrderSearchParams(self)
        self.render("product/productAdd.html", base_url=base_url,pager=pager,order_search_params=order_search_params)

    @gen.coroutine
    def post(self):
        product_name = self.get_argument('productName')
        class_name = self.get_argument('className')
        price = self.get_argument('price')
        other_name = self.get_argument('otherName')
        picture = self.get_argument('picture')
        ret = {'result': 'OK'}


        product_dic = dict(
            productName = product_name,
            className = class_name,
            otherName = other_name,
            price = price,
            picture = picture,
        )
        status = 'no'
        productinfo = self.async_do(HomeService.save_product, self.db, product_dic)
        if productinfo is not None :
            status = 'yes'
        else:
            status = 'no'

        ret['status'] = status
        respon_json = tornado.escape.json_encode(ret)  
        self.write_json(respon_json)

class ProductRemove(BaseHandler):
    @gen.coroutine
    def post(self):
        productid = self.get_argument('id')
        ret = {'result': 'OK'}
        status = 'false'
        productinfo = yield self.async_do(HomeService.remove_product, self.db, productid)
        if productinfo is not None :
            status = 'true'
        else:
            status = 'false'

        ret['status'] = status
        respon_json = tornado.escape.json_encode(ret)  
        self.write_json(respon_json)

class OrderCheckPrice(BaseHandler):
    @gen.coroutine
    def post(self):
        orderid = self.get_argument('id')
        checkPrice = float(self.get_argument('checkPrice'))
        ret = {'result': 'OK'}
        status = 'false'
        oldOrder = yield self.async_do(HomeService.get_order_by_id, self.db, orderid)
        if oldOrder is not None :
            status = 1
            if oldOrder.orderPrice_ > checkPrice:
                status = 2
            order_dic = dict(
                checkPrice_ = checkPrice,
                status_ = status,
            )
            order = yield self.async_do(HomeService.update_order, self.db, orderid, order_dic)
            if order is not None :
                status = 'true'
            else:
                status = 'no order not find'
        else:
            status = 'no oldOrder not find'

        ret['status'] = status
        respon_json = tornado.escape.json_encode(ret)  
        self.write_json(respon_json)

class FileUpload(BaseHandler):
    
    def get(self):
        pager = Pager(self)
        base_url = self.reverse_url('index')
        order_search_params = OrderSearchParams(self)
        self.render("order/index.html", base_url=base_url,pager=pager,order_search_params=order_search_params)

    @gen.coroutine
    def post(self):
        ret = {'result': 'OK'}
        upload_path = os.path.join(os.path.dirname(__file__), 'files')  # 文件的暂存路径
        file_metas = self.request.files.get('fileName', None)  # 提取表单中‘name’为‘file’的文件元数据

        if not file_metas:
            ret['result'] = 'Invalid Args'
            return ret

        for meta in file_metas:
            filename = meta['filename']
            file_path = os.path.join(upload_path, filename)
            with open(file_path, 'wb') as up:
                up.write(meta['body'])
            #     # OR do other thing

            wb = load_workbook(file_path,data_only=True)  #加载一个工作簿
            ret['name']=wb.get_sheet_names()
            #获取所有表格(worksheet)的名字
            sheets = wb.get_sheet_names()
            #第一个表格的名称
            sheet_first = sheets[0]
            #获取特定的worksheet
            ws = wb.get_sheet_by_name(sheet_first)
            #获取表格所有行和列，两者都是可迭代的
            # ws_rows_len = len(ws.rows)          #行数
            # ws_columns_len = len(ws.columns)    #列数
            # 建立存储数据的字典
            data_dic = {}

            status = 'xxxx'

            # 把数据存到字典中
            # for rx in range(2, ws.max_row + 1):
            for rx in range(2, ws.max_row):
            # for rx in range(2, 3):
                temp_list = []
                pid = rx
                for cs in range(1, ws.max_column + 1):
                    w1 = ws.cell(row=rx, column=cs).value
                    temp_list.append(w1)

                data_dic[pid] = temp_list
                print '名字：'+temp_list[2].replace(' ','').replace(' ','')
                order_dic = dict(
                    consignee = temp_list[2],
                    phone = temp_list[3],
                    provinceName = temp_list[5],
                    address = temp_list[8],
                    orderInfo = temp_list[10],
                    express = temp_list[11],
                    expressNumber = temp_list[12],
                    expressPrice = temp_list[21],
                    weight = temp_list[13],
                    printTime = temp_list[18],
                    userMark = temp_list[19],
                    orderPrice = temp_list[20],
                )
                # order_dic = {}
                # order_dic['consignee'] = temp_list[2]
                # order_dic['phone'] = temp_list[3]
                # order_dic['provinceName'] = temp_list[5]
                # order_dic['address'] = temp_list[8]
                # order_dic['orderInfo'] = temp_list[10]
                # order_dic['express'] = temp_list[11]
                # order_dic['expressNumber'] = temp_list[12]
                # order_dic['expressPrice'] = temp_list[21]
                # order_dic['weight'] = temp_list[13]
                # order_dic['printTime'] = temp_list[18]
                # order_dic['userMark'] = temp_list[19]
                # order_dic['orderPrice'] = temp_list[20]
                orderinfo = self.async_do(HomeService.save_order, self.db, order_dic)
                if orderinfo is not None :
                    status = 'yes'
                else:
                    status = 'no'


            self.db.commit()
            ret['line'] = data_dic
            ret['status'] = status

        respon_json = tornado.escape.json_encode(ret)  
        self.write(respon_json)

        
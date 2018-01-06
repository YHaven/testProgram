# coding=utf-8
import os
from tornado import gen
import tornado.escape 

from base import BaseHandler
from model.pager import Pager
from model.constants import Constants
from model.search_params.order_params import OrderSearchParams
from service.user_service import UserService
from service.home_service import HomeService
from openpyxl import load_workbook


class HomeHandler(BaseHandler):
    @gen.coroutine
    def get(self):
        pager = Pager(self)
        base_url = self.reverse_url('index')
        order_search_params = OrderSearchParams(self)
        self.render("index.html", base_url=base_url,pager=pager,order_search_params=order_search_params)


class LoginHandler(BaseHandler):

    def get(self):
        next_url = self.get_argument('next', '/')
        self.render("auth/login.html", next_url=next_url,pager=pager)

    @gen.coroutine
    def post(self):
        username = self.get_argument('username')
        password = self.get_argument('password')
        next_url = self.get_argument('next', '/')
        user = yield self.async_do(UserService.get_user, self.db, username)
        if user is not None and user.password == password:
            self.save_login_user(user)
            self.add_message('success', u'登陆成功！欢迎回来，{0}!'.format(username))
            self.redirect(next_url)
        else:
            self.add_message('danger', u'登陆失败！用户名或密码错误，请重新登陆。')
            self.get()


class LogoutHandler(BaseHandler):

    def get(self):
        self.logout()
        self.add_message('success', u'您已退出登陆。')
        self.redirect("/")


class FileUpload(BaseHandler):
    
    def get(self):
        pager = Pager(self)
        base_url = self.reverse_url('index')
        order_search_params = OrderSearchParams(self)
        self.render("index.html", base_url=base_url,pager=pager,order_search_params=order_search_params)

    @gen.coroutine
    def post(self):
        ret = {'result': 'OK'}
        upload_path = os.path.join(os.path.dirname(__file__), 'files')  # 文件的暂存路径
        file_metas = self.request.files.get('fileName', None)  # 提取表单中‘name’为‘file’的文件元数据

        if not file_metas:
            ret['result'] = 'Invalid Args'
            return ret

        for meta in file_metas:
            filename = meta['filename']
            file_path = os.path.join(upload_path, filename)
            with open(file_path, 'wb') as up:
                up.write(meta['body'])
            #     # OR do other thing

            wb = load_workbook(file_path,data_only=True)  #加载一个工作簿
            ret['name']=wb.get_sheet_names()
            #获取所有表格(worksheet)的名字
            sheets = wb.get_sheet_names()
            #第一个表格的名称
            sheet_first = sheets[0]
            #获取特定的worksheet
            ws = wb.get_sheet_by_name(sheet_first)
            #获取表格所有行和列，两者都是可迭代的
            # ws_rows_len = len(ws.rows)          #行数
            # ws_columns_len = len(ws.columns)    #列数
            # 建立存储数据的字典
            data_dic = {}

            status = 'xxxx'

            # 把数据存到字典中
            for rx in range(2, ws.max_row + 1):
            # for rx in range(2, 3):
                temp_list = []
                pid = rx
                for cs in range(1, ws.max_column + 1):
                    w1 = ws.cell(row=rx, column=cs).value
                    temp_list.append(w1)

                data_dic[pid] = temp_list
                order_dic = dict(
                    consignee = temp_list[2],
                    phone = temp_list[3],
                    provinceName = temp_list[5],
                    address = temp_list[8],
                    orderInfo = temp_list[10],
                    express = temp_list[11],
                    expressNumber = temp_list[12],
                    expressPrice = temp_list[21],
                    weight = temp_list[13],
                    printTime = temp_list[18],
                    userMark = temp_list[19],
                    orderPrice = temp_list[20],
                )
                # order_dic = {}
                # order_dic['consignee'] = temp_list[2]
                # order_dic['phone'] = temp_list[3]
                # order_dic['provinceName'] = temp_list[5]
                # order_dic['address'] = temp_list[8]
                # order_dic['orderInfo'] = temp_list[10]
                # order_dic['express'] = temp_list[11]
                # order_dic['expressNumber'] = temp_list[12]
                # order_dic['expressPrice'] = temp_list[21]
                # order_dic['weight'] = temp_list[13]
                # order_dic['printTime'] = temp_list[18]
                # order_dic['userMark'] = temp_list[19]
                # order_dic['orderPrice'] = temp_list[20]
                orderinfo = self.async_do(HomeService.save_order, self.db, order_dic)
                if orderinfo is not None :
                    status = 'yes'
                else:
                    status = 'no'


            self.db.commit()
            ret['line'] = data_dic
            ret['status'] = status

        respon_json = tornado.escape.json_encode(ret)  
        self.write(respon_json)